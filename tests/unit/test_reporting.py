# SPDX-FileCopyrightText: 2026 James G Willmore
# SPDX-License-Identifier: Apache-2.0

from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import openpyxl

from pydoublecross.reporting.exporters.excel import ExcelExporter
from pydoublecross.reporting.report import ge_result_rows, mismatch_rows, summary_rows
from pydoublecross.validation.results import (
    ColumnMismatch,
    ComparisonSummary,
    GESideResult,
    RunStatus,
    ValidationRunResult,
)


def _sample_result() -> ValidationRunResult:
    now = datetime.now(UTC)
    return ValidationRunResult(
        item_name="customer_check",
        run_id="abc123",
        started_at=now,
        finished_at=now,
        status=RunStatus.FAILED,
        source_cache_hit=False,
        target_cache_hit=True,
        summary=ComparisonSummary(
            source_row_count=3,
            target_row_count=3,
            matched_row_count=1,
            missing_in_target_count=1,
            missing_in_source_count=1,
            mismatched_row_count=1,
            mismatched_cell_count=1,
        ),
        missing_in_target=[{"customer_id": 3}],
        missing_in_source=[{"customer_id": 4}],
        mismatches=[
            ColumnMismatch(
                key={"customer_id": 2}, column="email", source_value="b@x.com", target_value="bad"
            )
        ],
        ge_results=[
            GESideResult(
                role="source", success=True, expectations_evaluated=3, expectations_failed=0
            ),
            GESideResult(
                role="target", success=True, expectations_evaluated=3, expectations_failed=0
            ),
        ],
    )


def test_summary_rows_includes_key_metrics() -> None:
    rows = summary_rows(_sample_result())
    metrics = {r["metric"] for r in rows}
    assert "status" in metrics
    assert "mismatched_cell_count" in metrics


def test_mismatch_rows_flattens_key() -> None:
    rows = mismatch_rows(_sample_result())
    assert rows == [
        {"key.customer_id": 2, "column": "email", "source_value": "b@x.com", "target_value": "bad"}
    ]


def test_ge_result_rows() -> None:
    rows = ge_result_rows(_sample_result())
    assert len(rows) == 2
    assert rows[0]["role"] == "source"


def test_excel_exporter_writes_all_sheets(tmp_path: Path) -> None:
    result = _sample_result()
    destination = tmp_path / "report"
    path = ExcelExporter().export(result, destination)
    assert path.suffix == ".xlsx"
    assert path.is_file()

    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {
        "Summary",
        "Missing In Target",
        "Missing In Source",
        "Value Mismatches",
        "GE Expectations",
    }
    assert wb["Value Mismatches"].max_row == 2  # header + 1 mismatch row


def test_excel_exporter_handles_empty_result(tmp_path: Path) -> None:
    now = datetime.now(UTC)
    result = ValidationRunResult(
        item_name="empty_check",
        run_id="empty1",
        started_at=now,
        finished_at=now,
        status=RunStatus.PASSED,
        summary=ComparisonSummary(
            source_row_count=0,
            target_row_count=0,
            matched_row_count=0,
            missing_in_target_count=0,
            missing_in_source_count=0,
            mismatched_row_count=0,
            mismatched_cell_count=0,
        ),
    )
    path = ExcelExporter().export(result, tmp_path / "empty")
    wb = openpyxl.load_workbook(path)
    assert wb["Value Mismatches"].max_row == 2  # header + the "no rows" note
